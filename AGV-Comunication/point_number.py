from database import Database, SQL
import config
import openpyxl
from pathlib import Path

def INSERT_POINT_NUMBER(filename):
    xlsx_file = Path('FILE', filename)
    wb_obj = openpyxl.load_workbook(xlsx_file) 
    
    # Read the active sheet:
    sheet = wb_obj.active
    add_row = []
    for row in sheet.iter_rows():
        for cell in row:
            if (str(cell.value)).upper() not in ['TASK', 'PATH', 'HOME', 'NONE']:
                add_row.append((str(cell.value)).upper())
                if len(add_row) == 3 and add_row[1] != '0':
                    print(add_row)
                    conn = SQL().connection()
                    db = Database(conn)
                    db.INSERT_POINT_NUMBER(int(add_row[0]), add_row[1], add_row[2])
                    add_row = []
            
def ListToStr(ListPoint):
    Str = ''
    for i in ListPoint:
        Str = Str + (',' + str(i)) if Str != '' else Str + str(i)
    return Str

def INSERT_SUPPLY(filename):
    xlsx_file = Path('FILE', filename)
    wb_obj = openpyxl.load_workbook(xlsx_file) 

    # Read the active sheet:
    sheet = wb_obj.active
    add_row = []
    for row in sheet.iter_rows():
        for cell in row:
            if (str(cell.value)).upper() not in ['SUPPLY', 'LINE', 'PROJECT', 'TASK', 'HOME', 'NONE']:
                add_row.append((str(cell.value)).upper())

                if len(add_row) == 5:
                    print(add_row)
                    conn = SQL().connection()
                    db = Database(conn)
                    db.INSERT_SUPPLY(add_row)
                    add_row = []
                    
def INSERT_POINT_NUMBER_CASE(filename):
    xlsx_file = Path('FILE', filename)
    wb_obj = openpyxl.load_workbook(xlsx_file) 
    
    # Read the active sheet:
    sheet = wb_obj.active
    add_row = []
    for row in sheet.iter_rows():
        for cell in row:
            if (str(cell.value)).upper() not in ['POINT', 'PATH_RANGE', 'HOME', 'NONE']:
                add_row.append((str(cell.value)).upper())
                if len(add_row) == 3:
                    print(add_row)
                    conn = SQL().connection()
                    db = Database(conn)
                    db.INSERT_POINT_NUMBER_CASE(int(add_row[0]), add_row[1], add_row[2])
                    add_row = []
                    
    
def INSERT_POINT_MAPPING(filename):
    xlsx_file = Path('FILE', filename)
    wb_obj = openpyxl.load_workbook(xlsx_file) 
    
    sheet = wb_obj.active
    add_row = []
    for row in sheet.iter_rows():
        for cell in row:
            if (str(cell.value)).upper() not in ['POINT_NUMBER', 'LEFT', 'TOP', 'NONE']:
                add_row.append((str(cell.value)).upper())
                if len(add_row) == 3:
                    print(add_row)
                    conn = SQL().connection()
                    db = Database(conn)
                    db.INSERT_POINT_MAPPING(int(add_row[0]), int(add_row[1]), int(add_row[2]))
                    add_row = []
            


supSI = list(range(1,15))
supCU = [15, 16, 17, 18]
supCT = list(range(21,31))

# ---------- Class CU-Store ----------------------------------------------
ROOM_CU = config.ROOM_CU

# insert path to CU-Store to point number
# INSERT_POINT_NUMBER('ROUTE-CU.xlsx')

# INSERT SUPPLY
# [SUPPLY, LINE, PROJECT, TASK, HOME]
# INSERT_SUPPLY('SUPPLY-STORE.xlsx')


# ---------- Class CU-Kitting ----------------------------------------------
ROOM_KWH = config.ROOM_KWH

# insert path to CU-Kitting to point number
# INSERT_POINT_NUMBER('ROUTE-KITTING.xlsx')

# insert point case
# INSERT_POINT_NUMBER_CASE('ROUTE-ALL-CASE.xlsx')

# INSERT SUPPLY
# [SUPPLY, LINE, PROJECT, TASK, HOME]
# INSERT_SUPPLY('SUPPLY-KITTING.xlsx')    # filename

# ---------- Class SIWH ----------------------------------------------------
ROOM_SI = config.ROOM_SI
# insert path to SI-WH to point number
# INSERT_POINT_NUMBER('ROUTE-SIWH.xlsx')

# INSERT SUPPLY
# [SUPPLY, LINE, PROJECT, TASK, HOME]
# INSERT_SUPPLY('xxxx.xlsx')
# --------------------------------------------------------------------------

# insert point mapping
# INSERT_POINT_MAPPING('mark_map.xlsx')